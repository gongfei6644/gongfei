# -*- coding: utf-8 -*-
# @Desc    : 导出标准化数据

import os
import shutil
import time
import zipfile

from openpyxl import load_workbook

from app.config import EXPORT_FILE_DIR
from app.excel import ExcelExporter
from app.models.std_case import *

logger = logging.getLogger(__name__)


class ExportStdCases(ExcelExporter):
    def __init__(self, city, start_date, end_date, status=[1], file=None):
        """
        自定义的excel读写器
        :param city: city_name
        :param start_date: case_happen_date
        :param end_date: case_happen_date
        :param status: todo
        :param file: todo
        """
        self.city = city
        self.start_date = start_date
        self.end_date = end_date
        self.status = status
        self.cust_file = file

    # 导出去重去偏差的标准化案例
    def export(self, page_index=1):
        # 复制模板
        file = self.copy_file()
        print('开始导出数据',page_index)
        file_paths = file
        # 导入数据到复制的文件中
        wb = load_workbook(file)
        sheet = wb['租金案例']
        # m_col = sheet.max_column
        is_exported = False
        row = 2
        while True:
            rt = self.get_list(page_index)
            if not rt:
                break
            for case in rt:
                self.assemble_row(sheet, case, row)
                row = row + 1
            page_index = page_index + 1
            # 一个excel数据太多会导致打开操作等很慢，所以设置了一个excel最多10万行
            # if row > 100000:
            if row > 100000:
                wb.save(file)
                is_exported = True
                ret = self.export(page_index)
                file_paths += ';' + ret[0]
                page_index = ret[1]

        if not is_exported:
            wb.save(file)
        return file_paths, page_index

    # 打包文件
    def zip(self, file_list):
        zip_file_name = '{}-{} {}租金标准化案例数据.zip'.format(
            self.start_date.replace('-', ''), self.end_date.replace('-', ''), self.city)
        zip_file = os.path.join(EXPORT_FILE_DIR, 'export-cases', 'renting_case', datetime.now().strftime('%Y-%m-%d'), zip_file_name)
        ret_path = zip_file
        zips = zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED)
        try:
            for path in file_list:
                zips.write(path, arcname=path[path.rfind(os.path.sep) + 1:])
        finally:
            zips.close()
        return ret_path

    def copy_file(self):
        """
        生成文件路径
        :return: 文件路径
        """
        if self.cust_file:
            dst_file = self.cust_file
        else:
            file = '{}-{} {}租金标准化案例数据_{}.xlsx'.format(
                self.start_date.replace('-', ''), self.end_date.replace('-', ''),
                self.city, str(time.time()).replace('.','')[-10:])
            dst_file = os.path.join(EXPORT_FILE_DIR, 'export-cases', 'renting_case', datetime.now().strftime('%Y-%m-%d'))
            if not os.path.exists(dst_file):
                os.makedirs(dst_file)
            dst_file = os.path.join(dst_file, file)

        _dir = os.path.dirname(os.path.dirname(__file__))
        src_file = os.path.join(_dir, 'renting_case_import_tpl.xlsx')
        shutil.copyfile(src_file, dst_file)
        return dst_file

    def get_list(self, page_index):
        rt = StdCase().get_list(self.city, self.start_date, self.end_date, page_index, status=self.status)
        return rt

    def assemble_row(self, sheet, case, row):
        self.set_cell(sheet, case.city_name, row, 1)
        self.set_cell(sheet, case.project_name, row, 2)
        self.set_cell(sheet, case.area_name, row, 3)
        self.set_cell(sheet, case.build_name, row, 4)
        self.set_cell(sheet, case.floor_no, row, 5)
        self.set_cell(sheet, case.house_name, row, 6)
        self.set_cell(sheet, case.total_floor_num, row, 7)
        self.set_cell(sheet, case.case_happen_date, row, 8)
        self.set_cell(sheet, case.usage, row, 9)
        self.set_cell(sheet, case.build_area, row, 10)
        self.set_cell(sheet, case.unitprice, row, 11)
        self.set_cell(sheet, case.total_price, row, 12)
        self.set_cell(sheet, case.case_type, row, 13)
        self.set_cell(sheet, case.rental_method, row, 14)
        self.set_cell(sheet, case.deposit_method, row, 15)
        self.set_cell(sheet, case.orientation, row, 16)
        self.set_cell(sheet, case.build_type, row, 17)
        self.set_cell(sheet, case.house_type, row, 18)
        self.set_cell(sheet, case.house_structure, row, 19)
        self.set_cell(sheet, case.build_date, row, 20)
        self.set_cell(sheet, case.decoration, row, 21)
        self.set_cell(sheet, case.usable_area, row, 22)
        self.set_cell(sheet, case.remaining_years, row, 23)
        self.set_cell(sheet, case.new_ratio, row, 24)
        self.set_cell(sheet, case.currency, row, 25)
        self.set_cell(sheet, case.affiliated_house, row, 26)
        self.set_cell(sheet, case.supporting_facilities, row, 27)  # 配套
        self.set_cell(sheet, case.data_source, row, 28)
        self.set_cell(sheet, case.source_link, row, 29)
        self.set_cell(sheet, case.tel, row, 30)  # 来源电话
        self.set_cell(sheet, case.remark, row, 31)  # 备注
