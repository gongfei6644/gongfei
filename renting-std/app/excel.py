# -*- coding: utf-8 -*-


from openpyxl import Workbook


class ExcelExporter:
    def __init__(self):
        self.wb = Workbook()

    def export(self, row0, file):
        sheet = self.wb.active
        for column in range(0, len(row0)):
            sheet.cell(row=1, column=column + 1).value = row0[column]
        page_index = 1
        row = 2
        while True:
            rt = self.get_list(page_index)
            if not rt:
                break
            for data in rt:
                self.assemble_row(sheet, data, row)
                row = row + 1
            page_index = page_index + 1
        self.wb.save(file)

    def get_list(self, page_index):
        print('')

    def assemble_row(self, sheet, data, row):
        print('')

    def set_cell(self, sheet, val, row, col):
        cell = sheet.cell(row=row, column=col)
        cell.value = val
